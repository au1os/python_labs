import csv
from pathlib import Path
try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    raise ImportError("Для работы модуля требуется openpyxl. Установите: pip install openpyxl")


def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
 
    csv_file = Path(csv_path)
    if not csv_file.exists():
        raise FileNotFoundError(f"CSV файл не найден: {csv_path}")
    
  
    with csv_file.open('r', encoding='utf-8') as f:
   
        sample = f.read(1024)
        f.seek(0)
        
        try:
            dialect = csv.Sniffer().sniff(sample)
            has_header = csv.Sniffer().has_header(sample)
        except:
            dialect = 'excel'
            has_header = True
        
        if not has_header:
            raise ValueError("CSV файл должен содержать заголовок")
        
        reader = csv.reader(f, dialect=dialect)
        rows = list(reader)
    

    if not rows:
        raise ValueError("CSV файл пустой")
    
  
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
  
    for row_idx, row in enumerate(rows, start=1):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
  
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        
        for cell in column_cells:
            try:
                cell_length = len(str(cell.value)) if cell.value else 0
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass
        
  
        adjusted_width = max(max_length + 2, 8)
        ws.column_dimensions[column_letter].width = adjusted_width
    
 
    xlsx_file = Path(xlsx_path)
    xlsx_file.parent.mkdir(parents=True, exist_ok=True)
    
 
    wb.save(xlsx_file)

def get_csv_info(csv_path: str) -> dict:

    with open(csv_path, 'r', encoding='utf-8') as f:
        sample = f.read(1024)
        f.seek(0)
        
        try:
            dialect = csv.Sniffer().sniff(sample)
            delimiter = dialect.delimiter
        except:
            delimiter = ','
        
        reader = csv.DictReader(f)
        rows = list(reader)
        
        return {
            'columns': reader.fieldnames or [],
            'row_count': len(rows),
            'delimiter': delimiter
        }